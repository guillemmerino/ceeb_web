# competicio/services/import_excel.py
import unicodedata
from datetime import datetime, date
from typing import Optional, Dict, Any, Set

from openpyxl import load_workbook

from ..models import Inscripcio, Competicio


def _norm_header(s: str) -> str:
    """
    Normalitza capçaleres: minúscules, sense accents, espais -> underscore.
    Ex: 'Data de naixement' -> 'data_de_naixement'
    """
    s = (s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.replace(" ", "_")
    s = s.replace("-", "_")
    s = s.replace("/", "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s


def _to_none(v):
    if v is None:
        return None
    if isinstance(v, str) and not v.strip():
        return None
    return v


def _parse_date(value) -> Optional[date]:
    """
    Accepta:
    - datetime/date d'Excel
    - string amb formats típics: dd/mm/yyyy o yyyy-mm-dd
    """
    value = _to_none(value)
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, str):
        txt = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(txt, fmt).date()
            except ValueError:
                pass

    return None


def importar_inscripcions_excel(fitxer, competicio: Competicio, sheet: str = "") -> Dict[str, Any]:
    """
    Importa inscripcions des del fitxer Excel adjunt.
    Mapeig (Excel -> model):
      - Nif -> document
      - Nom + Cognoms -> nom_i_cognoms
      - Entitat -> entitat
      - Categoria -> categoria
      - SubCategoria -> subcategoria
      - Sexe -> sexe
      - Data de naixement -> data_naixement

    Estratègia de duplicats:
      - Si hi ha Nif/document: update_or_create per (competicio, document)
      - Si no hi ha Nif: crea nou registre (no podem deduplicar de forma fiable)
    """
    wb = load_workbook(fitxer, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    # Índex de columnes a partir de la capçalera (fila 1)
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[_norm_header(str(cell.value))] = col_idx

    def get(row: int, *possible_headers: str):
        for h in possible_headers:
            idx = headers.get(_norm_header(h))
            if idx:
                return ws.cell(row=row, column=idx).value
        return None

    creats = 0
    actualitzats = 0
    ignorats = 0
    errors = 0
    noms_competicio_excel: Set[str] = set()

    for r in range(2, ws.max_row + 1):
        try:
            nif = _to_none(get(r, "Nif"))
            nom = _to_none(get(r, "Nom"))
            cognoms = _to_none(get(r, "Cognoms"))

            # Guardem el nom de competició del fitxer (per control/avisos)
            nom_comp_excel = _to_none(get(r, "Nom Competició"))
            if nom_comp_excel:
                noms_competicio_excel.add(str(nom_comp_excel).strip())

            # Sense nom/cognoms, no podem importar
            if not nom and not cognoms:
                ignorats += 1
                continue

            nom_i_cognoms = f"{str(nom).strip() if nom else ''} {str(cognoms).strip() if cognoms else ''}".strip()

            defaults = {
                "nom_i_cognoms": nom_i_cognoms,
                "entitat": _to_none(get(r, "Entitat")),
                "categoria": _to_none(get(r, "Categoria")),
                "subcategoria": _to_none(get(r, "SubCategoria")),
                "sexe": _to_none(get(r, "Sexe")),
                "data_naixement": _parse_date(get(r, "Data de naixement")),
            }

            # Normalitzacions suaus
            if defaults["sexe"] is not None:
                defaults["sexe"] = str(defaults["sexe"]).strip()

            if defaults["categoria"] is not None:
                defaults["categoria"] = str(defaults["categoria"]).strip()

            if defaults["subcategoria"] is not None:
                defaults["subcategoria"] = str(defaults["subcategoria"]).strip()

            if defaults["entitat"] is not None:
                defaults["entitat"] = str(defaults["entitat"]).strip()

            if nif:
                document = str(nif).strip()
                obj, created = Inscripcio.objects.update_or_create(
                    competicio=competicio,
                    document=document,
                    defaults=defaults,
                )
                if created:
                    creats += 1
                else:
                    actualitzats += 1
            else:
                Inscripcio.objects.create(
                    competicio=competicio,
                    document=None,
                    **defaults,
                )
                creats += 1

        except Exception:
            errors += 1

    return {
        "full": ws.title,
        "creats": creats,
        "actualitzats": actualitzats,
        "ignorats": ignorats,
        "errors": errors,
        "noms_competicio_excel": sorted(noms_competicio_excel),
    }
