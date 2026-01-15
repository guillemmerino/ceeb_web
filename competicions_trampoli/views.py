import random
from django.shortcuts import render
import math
from django.contrib import messages
from django.shortcuts import get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import CreateView, FormView, ListView, TemplateView, DeleteView
from django.shortcuts import redirect
from .forms import ImportInscripcionsExcelForm
from .models import Competicio, Inscripcio
from .forms import CompeticioForm
from .services.import_excel import importar_inscripcions_excel
import json
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from django.db import transaction
from django.views.generic import UpdateView
from django.urls import reverse
from .forms import InscripcioForm
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models import Max, Min, Case, When, IntegerField



ALLOWED_SORT_FIELDS = {
    "nom": "nom_i_cognoms",
    "edat": "data_naixement",        # per edat, ordenarem per data naixement
    "document": "document",
    "sexe": "sexe",
    "entitat": "entitat__nom",
    "categoria": "categoria__nom",
    "subcategoria": "subcategoria__nom",
    "grup": "grup",
}

ALLOWED_GROUP_FIELDS = {
    "categoria": "categoria",
    "subcategoria": "subcategoria",
    "entitat": "entitat",
    "sexe": "sexe",
    # si més endavant vols: "prova": "prova"
}

def _s(v):
    return "(Sense valor)" if v in (None, "") else str(v)

def arrow_positions(n: int) -> list[int]:
    """
    Retorna la seqüència de posicions "fletxa" per un grup de mida n.

    Exemple n=8 -> [3,4,2,5,1,6,0,7]
    (index del registre ordenat) -> (posició dins del grup)
    """
    if n <= 0:
        return []

    seq = []
    if n % 2 == 0:
        left = n // 2 - 1
        right = n // 2
        while left >= 0 or right < n:
            if left >= 0:
                seq.append(left)
                left -= 1
            if right < n:
                seq.append(right)
                right += 1
    else:
        center = n // 2
        seq.append(center)
        step = 1
        while center - step >= 0 or center + step < n:
            if center - step >= 0:
                seq.append(center - step)
            if center + step < n:
                seq.append(center + step)
            step += 1

    return seq

def init_ordre_sortida(competicio_id):
    qs = Inscripcio.objects.filter(competicio_id=competicio_id).order_by("id")
    with transaction.atomic():
        for i, obj in enumerate(qs, start=1):
            Inscripcio.objects.filter(id=obj.id).update(ordre_sortida=i)

def shuffle_ordre_sortida(qs):
    ids = list(qs.values_list("id", flat=True))
    random.shuffle(ids)

    with transaction.atomic():
        for idx, ins_id in enumerate(ids, start=1):
            Inscripcio.objects.filter(id=ins_id).update(ordre_sortida=idx)



UNDO_SESSION_KEY = "inscripcions_undo_state"

def save_undo_state(request, qs):
    """
    Desa a sessió l'estat actual (grup + ordre_sortida)
    del queryset passat.
    """
    request.session[UNDO_SESSION_KEY] = list(
        qs.values("id", "grup", "ordre_sortida")
    )
    request.session.modified = True


def restore_undo_state(request):
    """
    Restaura l'últim estat guardat.
    Retorna nombre de registres restaurats.
    """
    data = request.session.get(UNDO_SESSION_KEY)
    if not data:
        return 0

    with transaction.atomic():
        for row in data:
            Inscripcio.objects.filter(id=row["id"]).update(
                grup=row["grup"],
                ordre_sortida=row["ordre_sortida"],
            )

    del request.session[UNDO_SESSION_KEY]
    request.session.modified = True
    return len(data)

def assign_groups_balanced(objs, size, start_group_num):
    """
    Reparteix objs en k grups, on k = ceil(n/size),
    i distribueix mides perquè difereixin com a màxim 1.
    Retorna el nou start_group_num (últim grup assignat).
    """
    n = len(objs)
    if n == 0:
        return start_group_num

    k = math.ceil(n / size)           # nombre de grups
    base = n // k                     # mida base de cada grup
    rem = n % k                       # els primers 'rem' grups tindran +1

    idx = 0
    group_num = start_group_num
    for g in range(k):
        group_num += 1
        this_size = base + (1 if g < rem else 0)
        for _ in range(this_size):
            objs[idx].grup = group_num
            idx += 1

    return group_num

def renumber_groups_for_competicio(competicio):
    """
    Re-numera grups consecutivament 1..N dins la competició, evitant forats.
    L'ordre de renumeració segueix la seva aparició a la llista (min ordre_sortida).
    """
    base = Inscripcio.objects.filter(competicio=competicio, grup__isnull=False)

    groups = list(
        base.values("grup")
            .annotate(min_ord=Min("ordre_sortida"))
            .order_by("min_ord", "grup")
    )
    if not groups:
        return

    mapping = {g["grup"]: i + 1 for i, g in enumerate(groups)}

    whens = [When(grup=old, then=new) for old, new in mapping.items()]
    base.update(
        grup=Case(
            *whens,
            default=None,
            output_field=IntegerField(),
        )
    )



class CompeticioHomeView(TemplateView):
    template_name = "competicio/home.html"


class CompeticioCreateView(CreateView):
    model = Competicio
    form_class = CompeticioForm
    template_name = "competicio/competicio_form.html"
    success_url = reverse_lazy("created")


class CompeticioDeleteView(DeleteView):
    model = Competicio
    template_name = "competicio/competicio_confirm_delete.html"
    success_url = reverse_lazy("created")

class InscripcionsImportExcelView(FormView):
    template_name = "competicio/inscripcions_import.html"
    form_class = ImportInscripcionsExcelForm

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["competicio"] = self.competicio
        return ctx

    def form_valid(self, form):
        fitxer = form.cleaned_data["fitxer"]
        sheet = form.cleaned_data.get("sheet") or ""

        result = importar_inscripcions_excel(fitxer, self.competicio, sheet)
        messages.success(
            self.request,
            f"Importació OK. Full: {result['full']} | Creats: {result['creats']} | "
            f"Actualitzats: {result['actualitzats']} | Ignorats: {result['ignorats']}"
        )
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("inscripcions_list", kwargs={"pk": self.competicio.pk})


def recalcular_ordre_sortida(qs, group_fields):
    # Primari: camps de grup (perquè quedin agrupats)
    # Secundari: ordre_sortida (manté l'ordre previ, que pot ser aleatori)
    order = [f for f in group_fields if f] + ["ordre_sortida", "id"]
    ids = list(qs.order_by(*order).values_list("id", flat=True))

    with transaction.atomic():
        for idx, ins_id in enumerate(ids, start=1):
            Inscripcio.objects.filter(id=ins_id).update(ordre_sortida=idx)

class InscripcionsListView(ListView):
    template_name = "competicio/inscripcions_list.html"
    context_object_name = "records"
    paginate_by = 25

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_queryset_base_filtrada(self):
        qs = Inscripcio.objects.filter(competicio=self.competicio)

        q = self.request.GET.get("q")
        categoria = self.request.GET.get("categoria")
        prova = self.request.GET.get("prova")
        subcategoria = self.request.GET.get("subcategoria")
        entitat = self.request.GET.get("entitat")

        if subcategoria:
            qs = qs.filter(subcategoria__iexact=subcategoria)
        if entitat:
            qs = qs.filter(entitat__icontains=entitat)
        if q:
            qs = qs.filter(nom_i_cognoms__icontains=q)
        if categoria:
            qs = qs.filter(categoria__iexact=categoria)
        if prova:
            qs = qs.filter(prova__iexact=prova)

        return qs

    def get(self, request, *args, **kwargs):
        # 0z) Desfer última acció
        if request.GET.get("undo") == "1":
            restored = restore_undo_state(request)

            if restored:
                messages.success(request, f"S'ha desfet l'última acció ({restored} inscripcions).")
            else:
                messages.info(request, "No hi ha cap acció per desfer.")

            query = request.GET.copy()
            query.pop("undo", None)
            return redirect(f"{request.path}?{query.urlencode()}")


        # 0x) Exportar Excel (respectant ordre_sortida actual)
        if request.GET.get("export_excel") == "1":
            # Camps que vols al títol del grup (només els que també es poden usar per agrupar)
            # Ex: title_fields=categoria&title_fields=subcategoria
            title_fields_keys = [k for k in request.GET.getlist("title_fields") if k in ALLOWED_GROUP_FIELDS]
            title_fields = [ALLOWED_GROUP_FIELDS[k] for k in title_fields_keys]

            qs = self.get_queryset_base_filtrada().order_by("ordre_sortida", "id")
            save_undo_state(request, qs)

            # Camps que exportem al full (pots ajustar-los)
            columns = [
                ("Ordre", "ordre_sortida"),
                ("Grup", "grup"),
                ("Nom i cognoms", "nom_i_cognoms"),
                ("Categoria", "categoria"),
                ("Subcategoria", "subcategoria"),
                ("Entitat", "entitat"),
                ("Sexe", "sexe"),
                ("Data naixement", "data_naixement"),
                ("Document", "document"),
            ]

            wb = Workbook()
            ws = wb.active
            ws.title = "Inscripcions"

            bold = Font(bold=True)
            title_font = Font(bold=True, size=12)
            center = Alignment(vertical="center")
            group_fill = PatternFill("solid", fgColor="DBEAFE")  # blau suau

            row = 1

            # --- Capçalera amb competició i filtres actius ---
            ws.cell(row=row, column=1, value="Competició:").font = bold
            ws.cell(row=row, column=2, value=self.competicio.nom)
            row += 1

            # Mostrem filtres actius (q/categoria/subcategoria/entitat/prova)
            active_filters = []
            for key in ["q", "categoria", "subcategoria", "entitat", "prova"]:
                v = request.GET.get(key)
                if v:
                    active_filters.append(f"{key}={v}")

            ws.cell(row=row, column=1, value="Filtres:").font = bold
            ws.cell(row=row, column=2, value=", ".join(active_filters) if active_filters else "(cap)")
            row += 2

            # --- Sense grups? -> export plana ---
            has_groups = qs.exclude(grup__isnull=True).exists()

            def write_table_header(r):
                for col_idx, (label, _) in enumerate(columns, start=1):
                    c = ws.cell(row=r, column=col_idx, value=label)
                    c.font = bold
                    c.alignment = center

            def write_row(r, obj):
                for col_idx, (_, field) in enumerate(columns, start=1):
                    v = getattr(obj, field)
                    ws.cell(row=r, column=col_idx, value=v)

            if not has_groups:
                write_table_header(row)
                row += 1
                for obj in qs:
                    write_row(row, obj)
                    row += 1
            else:
                # --- Export per blocs de grup (respectant ordre_sortida) ---
                current_group = None
                buffer = []

                def flush_group(buf):
                    nonlocal row, current_group
                    if not buf:
                        return

                    # títol del grup: "<camp1> <camp2> ... Grup X"
                    sample = buf[0]
                    title_parts = []
                    for f in title_fields:
                        title_parts.append(_s(getattr(sample, f)))

                    grp_num = getattr(sample, "grup")
                    if grp_num is None:
                        group_title = "Sense grup"
                    else:
                        group_title = (" ".join([p for p in title_parts if p])) + f" Grup {grp_num}"
                        group_title = group_title.strip()

                    # fila títol (merge)
                    ws.cell(row=row, column=1, value=group_title).font = title_font
                    ws.cell(row=row, column=1).fill = group_fill
                    ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
                    row += 1

                    # capçalera taula
                    write_table_header(row)
                    row += 1

                    # dades
                    for o in buf:
                        write_row(row, o)
                        row += 1

                    row += 1  # espai

                for obj in qs:
                    g = obj.grup
                    if current_group is None:
                        current_group = g
                    if g != current_group:
                        flush_group(buffer)
                        buffer = []
                        current_group = g
                    buffer.append(obj)

                flush_group(buffer)

            # Ajust d'amplades (simple)
            for i, (label, _) in enumerate(columns, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(12, min(35, len(label) + 4))

            out = BytesIO()
            wb.save(out)
            out.seek(0)

            filename = f"inscripcions_competicio_{self.competicio.pk}.xlsx"
            resp = HttpResponse(
                out.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            return resp



        # 0) Esborrar grups assignats (sempre ha de funcionar)
        if request.GET.get("clear_groups") == "1":
            qs = self.get_queryset_base_filtrada()  # o tots si vols
            save_undo_state(request, qs)

            with transaction.atomic():
                qs.update(grup=None)
                renumber_groups_for_competicio(self.competicio)

            query = request.GET.copy()
            query.pop("clear_groups", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        # 0c) Fer grup independent per un subgrup (header)
        if request.GET.get("make_independent_group") == "1":
            lvl = request.GET.get("lvl")  # "g1", "g2", "g3"

            # camps seleccionats (per saber què és g1/g2/g3)
            selected = self.request.GET.getlist("group_by")
            if not selected:
                selected = self.competicio.group_by_default or []
            group_fields = [ALLOWED_GROUP_FIELDS[g] for g in selected if g in ALLOWED_GROUP_FIELDS]

            # Validació mínima
            if not group_fields:
                messages.error(request, "No hi ha agrupació activa per poder crear un grup independent.")
                query = request.GET.copy()
                query.pop("make_independent_group", None)
                query.pop("lvl", None)
                query.pop("v1", None)
                query.pop("v2", None)
                query.pop("v3", None)
                return redirect(f"{request.path}?{query.urlencode()}")

            # Valors del header (passem v1/v2/v3)
            v1 = request.GET.get("v1")
            v2 = request.GET.get("v2")
            v3 = request.GET.get("v3")

            # Quants nivells aplicar segons header clicat
            if lvl == "g1":
                upto = 1
            elif lvl == "g2":
                upto = 2
            elif lvl == "g3":
                upto = 3
            else:
                upto = 1

            upto = min(upto, len(group_fields))

            # Construïm el filtre del subgrup
            vals = [v1, v2, v3][:upto]
            filtres = {}

            # Per diferenciar "sense valor", usem el sentinel "__NULL__"
            for f, v in zip(group_fields[:upto], vals):
                if v == "__NULL__":
                    filtres[f"{f}__isnull"] = True
                else:
                    filtres[f] = v

            # QS del subgrup dins els filtres actuals (q/categoria/subcategoria/entitat...)
            sub_qs = self.get_queryset_base_filtrada().filter(**filtres)
            save_undo_state(request, self.get_queryset_base_filtrada())



            # 1) Si el subgrup ja té grup(s), fem servir el mínim
            existing_groups = list(
                sub_qs.exclude(grup__isnull=True).values_list("grup", flat=True).distinct()
            )

            if existing_groups:
                new_group_num = min(existing_groups)

                with transaction.atomic():
                    # Allibera aquest número de grup dins la competició per evitar duplicats
                    Inscripcio.objects.filter(competicio=self.competicio, grup=new_group_num).update(grup=None)
                    updated = sub_qs.update(grup=new_group_num)
                    renumber_groups_for_competicio(self.competicio)

            else:
                # 2) Si no hi havia grups previs al subgrup, mantenim lògica actual: max + 1
                max_grup = (
                    Inscripcio.objects
                    .filter(competicio=self.competicio)
                    .aggregate(m=Max("grup"))["m"]
                    or 0
                )
                new_group_num = max_grup + 1

                with transaction.atomic():
                    updated = sub_qs.update(grup=new_group_num)
                    renumber_groups_for_competicio(self.competicio)


            messages.success(request, f"Creat el grup {new_group_num} amb {updated} inscripcions del subgrup.")
            # Si és una petició AJAX, retornem JSON i NO redirigim
            is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
            if is_ajax:
                return JsonResponse({"ok": True, "new_group_num": new_group_num, "updated": updated})
            query = request.GET.copy()
            query.pop("make_independent_group", None)
            query.pop("lvl", None)
            query.pop("v1", None)
            query.pop("v2", None)
            query.pop("v3", None)
            return redirect(f"{request.path}?{query.urlencode()}")


        # 0d) Reendreçar dins de cada grup segons un camp
        if request.GET.get("sort_within_groups") == "1":
            sort_key = request.GET.get("sort_key") or "nom"
            sort_dir = request.GET.get("sort_dir") or "asc"  # asc / desc
            if sort_dir not in ("asc", "desc", "arrow_asc", "arrow_desc"):
                sort_dir = "asc"
            if sort_key not in ALLOWED_SORT_FIELDS:
                messages.error(request, "Camp d'ordenació no vàlid.")
                query = request.GET.copy()
                query.pop("sort_within_groups", None)
                query.pop("sort_key", None)
                query.pop("sort_dir", None)
                return redirect(f"{request.path}?{query.urlencode()}")

            order_field = ALLOWED_SORT_FIELDS[sort_key]

            qs = self.get_queryset_base_filtrada()
            save_undo_state(request, qs)


            if not qs.exclude(grup__isnull=True).exists():
                messages.info(request, "No hi ha grups assignats per poder reendreçar dins de grup.")
                query = request.GET.copy()
                query.pop("sort_within_groups", None)
                query.pop("sort_key", None)
                query.pop("sort_dir", None)
                return redirect(f"{request.path}?{query.urlencode()}")

            updates = []
            idx = 1

            # Grups presents (dins el que estàs veient)
            group_nums = list(
                qs.exclude(grup__isnull=True)
                .order_by("grup")
                .values_list("grup", flat=True)
                .distinct()
            )

            with transaction.atomic():
                for g in group_nums:
                    group_qs = qs.filter(grup=g)

                    if sort_dir in ("arrow_asc", "arrow_desc"):
                        base_prefix = "-" if sort_dir == "arrow_desc" else ""
                        # 1) Base order (asc) pel camp triat
                        base_list = list(group_qs.order_by(f"{base_prefix}{order_field}", "id").only("id"))

                        n = len(base_list)
                        pos = arrow_positions(n)  # index del registre ordenat -> posició al grup

                        placed = [None] * n
                        for i, obj in enumerate(base_list):
                            placed[pos[i]] = obj

                        # 2) Assignem ordre_sortida seguint la posició final (0..n-1)
                        for obj in placed:
                            obj.ordre_sortida = idx
                            updates.append(obj)
                            idx += 1

                    else:
                        # asc / desc normal
                        prefix = "-" if sort_dir == "desc" else ""
                        base_list = list(group_qs.order_by(f"{prefix}{order_field}", "id").only("id"))
                        for obj in base_list:
                            obj.ordre_sortida = idx
                            updates.append(obj)
                            idx += 1

                Inscripcio.objects.bulk_update(updates, ["ordre_sortida"], batch_size=500)



            messages.success(request, f"Reendreçat dins dels grups per '{sort_key}' ({sort_dir}).")

            query = request.GET.copy()
            query.pop("sort_within_groups", None)
            query.pop("sort_key", None)
            query.pop("sort_dir", None)
            return redirect(f"{request.path}?{query.urlencode()}")


        # 0b) Crear grups de mida N (respectant group_by però amb numeració global única)
        if request.GET.get("make_groups") == "1":
            try:
                size = int(request.GET.get("group_size") or 0)
            except ValueError:
                size = 0

            if size < 2:
                messages.error(request, "La mida del grup ha de ser com a mínim 2.")
                query = request.GET.copy()
                query.pop("make_groups", None)
                return redirect(f"{request.path}?{query.urlencode()}")

            group_mode = request.GET.get("group_mode") or "fixed"  # "fixed" (actual) o "balanced"

            # group_by seleccionat (ex: ["categoria","subcategoria"])
            group_by_keys = [g for g in request.GET.getlist("group_by") if g in ALLOWED_GROUP_FIELDS]
            group_fields = [ALLOWED_GROUP_FIELDS[g] for g in group_by_keys]  # camps reals

            # Ordenem per subgrup + ordre_sortida per mantenir el teu ordre previ dins cada subgrup
            qs = self.get_queryset_base_filtrada().order_by(*group_fields, "ordre_sortida", "id")
            save_undo_state(request, qs)

            objs = list(qs.only("id", "grup", *group_fields))

            global_group_num = 0

            if not group_fields:
                if group_mode == "balanced":
                    global_group_num = assign_groups_balanced(objs, size, global_group_num)
                else:
                    # mode actual (fixed)
                    for idx, obj in enumerate(objs, start=1):
                        obj.grup = (idx - 1) // size + 1

            else:
                if group_mode == "balanced":
                    current_key = None
                    buffer = []

                    def flush_buffer():
                        nonlocal global_group_num, buffer
                        if buffer:
                            global_group_num = assign_groups_balanced(buffer, size, global_group_num)
                            buffer = []

                    for obj in objs:
                        key = tuple(getattr(obj, f) for f in group_fields)

                        if key != current_key:
                            flush_buffer()
                            current_key = key

                        buffer.append(obj)

                    flush_buffer()

                else:
                    # mode actual (fixed)
                    current_key = None
                    count_in_chunk = 0

                    for obj in objs:
                        key = tuple(getattr(obj, f) for f in group_fields)

                        if key != current_key:
                            current_key = key
                            count_in_chunk = 0

                        if count_in_chunk == 0:
                            global_group_num += 1

                        obj.grup = global_group_num
                        count_in_chunk += 1

                        if count_in_chunk >= size:
                            count_in_chunk = 0


            with transaction.atomic():
                Inscripcio.objects.bulk_update(objs, ["grup"])
                renumber_groups_for_competicio(self.competicio)

            query = request.GET.copy()
            query.pop("make_groups", None)
            return redirect(f"{request.path}?{query.urlencode()}")


        # 1) Treure agrupació explícitament
        if request.GET.get("clear_group") == "1":
            self.competicio.group_by_default = []
            self.competicio.save(update_fields=["group_by_default"])

            query = request.GET.copy()
            query.pop("clear_group", None)
            # assegura que no queda cap group_by
            query.setlist("group_by", [])
            return redirect(f"{request.path}?{query.urlencode()}")

        # 2) Recalcular ordre segons agrupació actual
        if request.GET.get("recalc_order") == "1":
            group_fields = [
                ALLOWED_GROUP_FIELDS[g]
                for g in request.GET.getlist("group_by")
                if g in ALLOWED_GROUP_FIELDS
            ]
            qs = self.get_queryset_base_filtrada()
            save_undo_state(request, qs)

            recalcular_ordre_sortida(qs, group_fields)
            renumber_groups_for_competicio(self.competicio)

            query = request.GET.copy()
            query.pop("recalc_order", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        # 3) Shuffle (barreja aleatòria)
        if request.GET.get("shuffle_order") == "1":
            qs = self.get_queryset_base_filtrada()
            save_undo_state(request, qs)

            shuffle_ordre_sortida(qs)

            query = request.GET.copy()
            query.pop("shuffle_order", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        # 4) Persistència de group_by
        if "group_by" in request.GET:
            selected = [g for g in request.GET.getlist("group_by") if g in ALLOWED_GROUP_FIELDS]
            if selected != (self.competicio.group_by_default or []):
                self.competicio.group_by_default = selected
                self.competicio.save(update_fields=["group_by_default"])
        else:
            # 5) Aplicar el guardat quan no ve group_by
            saved = self.competicio.group_by_default or []
            if saved:
                query = request.GET.copy()
                for g in saved:
                    query.appendlist("group_by", g)
                return redirect(f"{request.path}?{query.urlencode()}")

        return super().get(request, *args, **kwargs)



    def get_queryset(self):
        qs = self.get_queryset_base_filtrada()
        save_undo_state(self.request, qs)

        return qs.order_by("ordre_sortida", "id")

    def get_paginate_by(self, queryset):
        # mantinc la teva lògica
        if self.request.GET.getlist("group_by"):
            return None
        return int(self.request.GET.get("per_page") or 10)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["competicio"] = self.competicio
        ctx["allowed_group_fields"] = ["categoria", "subcategoria", "entitat"]
        ctx["current_query"] = self.request.GET.urlencode()
        ctx["allowed_group_fields"] = list(ALLOWED_GROUP_FIELDS.keys())
        ctx["title_fields_selected"] = self.request.GET.getlist("title_fields")

        selected = self.request.GET.getlist("group_by")
        if not selected:
            selected = self.competicio.group_by_default or []

        ctx["selected_group_fields"] = selected
        return ctx


class InscripcioUpdateView(UpdateView):
    model = Inscripcio
    pk_url_kwarg = "ins_id"
    form_class = InscripcioForm
    template_name = "competicio/inscripcio_form.html"

    def get_queryset(self):
        # Seguretat: només permet editar inscripcions de la competició del pk
        return Inscripcio.objects.filter(competicio_id=self.kwargs["pk"])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["competicio"] = get_object_or_404(Competicio, pk=self.kwargs["pk"])
        ctx["next"] = self.request.GET.get("next", "")
        return ctx

    def get_success_url(self):
        nxt = self.request.GET.get("next")
        if nxt:
            return nxt
        return reverse("inscripcions_list", kwargs={"pk": self.kwargs["pk"]})


class InscripcioCreateView(CreateView):
    model = Inscripcio
    form_class = InscripcioForm
    template_name = "competicio/inscripcio_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["competicio"] = self.competicio
        ctx["next"] = self.request.GET.get("next", "")
        return ctx

    def form_valid(self, form):
        # Assign the competition before saving
        form.instance.competicio = self.competicio
        # If ordre_sortida not provided, append to the end
        if not form.instance.ordre_sortida:
            max_ord = Inscripcio.objects.filter(competicio=self.competicio).aggregate(m=Max("ordre_sortida"))["m"] or 0
            form.instance.ordre_sortida = max_ord + 1
        return super().form_valid(form)

    def get_success_url(self):
        nxt = self.request.GET.get("next")
        if nxt:
            return nxt
        return reverse("inscripcions_list", kwargs={"pk": self.kwargs["pk"]})


class InscripcioDeleteView(DeleteView):
    model = Inscripcio
    pk_url_kwarg = "ins_id"
    template_name = "competicio/inscripcio_confirm_delete.html"

    def get_queryset(self):
        return Inscripcio.objects.filter(competicio_id=self.kwargs["pk"])

    def get_success_url(self):
        nxt = self.request.GET.get("next")
        if nxt:
            return nxt
        return reverse("inscripcions_list", kwargs={"pk": self.kwargs["pk"]})


@require_POST
@csrf_protect
def inscripcions_reorder(request, pk):
    """
    Rep:
      {
        "ids": [<id1>, <id2>, ...],   # ordre final després del drag
        "moved_id": <id>,            # el registre arrossegat
        "new_index": <int>           # posició nova (0-based) dins ids
      }

    Guarda ordre_sortida = 1..N
    I (NOU) només pel registre mogut: adopta el grup del registre immediatament superior.
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
        ids = payload.get("ids", [])
        moved_id = payload.get("moved_id", None)
        new_index = payload.get("new_index", None)

        if not isinstance(ids, list) or not ids:
            return HttpResponseBadRequest("Payload invàlid")
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    wanted = [int(x) for x in ids if str(x).isdigit()]
    if not wanted:
        return HttpResponseBadRequest("IDs buits")

    # Validació moved_id / new_index (opcional però recomanada)
    if moved_id is not None:
        try:
            moved_id = int(moved_id)
        except Exception:
            return HttpResponseBadRequest("moved_id invàlid")

    if new_index is not None:
        try:
            new_index = int(new_index)
        except Exception:
            return HttpResponseBadRequest("new_index invàlid")

    # Ens assegurem que només reordenem inscripcions d'aquesta competició
    qs = Inscripcio.objects.filter(competicio_id=pk, id__in=wanted)
    found = set(qs.values_list("id", flat=True))
    if set(wanted) != found:
        return HttpResponseBadRequest("IDs no vàlids per aquesta competició")

    # Diccionari id->grup (estat actual abans de canviar)
    id_to_group = dict(qs.values_list("id", "grup"))

    with transaction.atomic():
        # 1) actualitza ordre_sortida
        for idx, ins_id in enumerate(wanted, start=1):
            Inscripcio.objects.filter(id=ins_id).update(ordre_sortida=idx)

        # 2) (NOU) només el registre mogut adopta el grup del superior immediat
        if moved_id is not None and new_index is not None and moved_id in wanted:
            if new_index > 0:
                prev_id = wanted[new_index - 1]
                prev_group = id_to_group.get(prev_id)

                # Mateixa nota: si NO vols que None esborri el grup, fes:
                # if prev_group is not None:
                Inscripcio.objects.filter(id=moved_id).update(grup=prev_group)

    return JsonResponse({"ok": True})



class CompeticioListView(ListView):
    model = Competicio
    template_name = "competicio/competicio_created_list.html"
    context_object_name = "competicions"
    paginate_by = 20